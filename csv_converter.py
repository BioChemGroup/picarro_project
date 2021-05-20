import csv
import openpyxl as px
import math

"""
同一フォルダ内にあるcsvファイルを読み取り、york法に必要なデータだけをエクセルに表示させるプログラム

操作手順
①フォルダ内のcsv名（拡張子不要）の入力
②新しく作成するexcel名（拡張子不要）の入力
"""

book = px.Workbook()
wc = book.worksheets[0]
sheet = book.active
sheet.title = "calculation"
i = 2

wc.cell(1, 1, "H2O濃度")
wc.cell(1, 2, "濃度逆数")
wc.cell(1, 3, "δ17O")
wc.cell(1, 4, "δ18O")
wc.cell(1, 5, "Δ17O")
wc.cell(1, 6, "濃度逆数分散")
wc.cell(1, 7, "Δ17O分散")
wc.cell(1, 8, "δ18O分散")

temp_csv = str(input("csvファイル名を入力："))

with open(temp_csv + ".csv") as f:
    for row in csv.reader(f):
        value0 = f"{row[0]}"
        value17 = f"{row[4]}"
        value18 = f"{row[5]}"
        valueH2O = f"{row[8]}"
        error18 = f"{row[14]}"
        error_cap17 = f"{row[16]}"
        errorH2O = f"{row[17]}"
        try:
            float_value0 = float(value0)
            float_value17 = float(value17)
            float_value18 = float(value18)
            value_cap17 = 1000 * (math.log(1 + (0.001 * float_value17)) - 0.528 * math.log(1 + (0.001 * float_value18)))
            float_valueH2O = float(valueH2O)
            var18 = float(error18)**2
            var_cap17 = float(error_cap17)**2
            sdH2O_puls = 1/(float_valueH2O + float(errorH2O))
            sdH2O_minus = 1/(float_valueH2O - float(errorH2O))
            varH2O = ((sdH2O_minus - sdH2O_puls) / 2)**2
            wc.cell(i, 1, float_valueH2O)
            wc.cell(i, 2, "=1/A" + str(i))
            wc.cell(i, 3, float_value17)
            wc.cell(i, 4, float_value18)
            wc.cell(i, 5, value_cap17)
            wc.cell(i, 6, varH2O)
            wc.cell(i, 7, var_cap17)
            wc.cell(i, 8, var18)
            i += 1
        except ValueError:
            pass

i -= 1
# δ18Oグラフの作成
chart18 = px.chart.ScatterChart()
x18 = px.chart.Reference(sheet, min_col=2, max_col=2, min_row=2, max_row=i)
y18 = px.chart.Reference(sheet, min_col=4, max_col=4, min_row=2, max_row=i)
series18 = px.chart.Series(y18, x18)
series18.graphicalProperties.line.noFill = True
series18.marker.symbol = "auto"
chart18.title = "δ18O"
chart18.x_axis.title = "1/濃度"
chart18.y_axis.title = "δ18O"
chart18.legend = None
chart18.series.append(series18)
sheet.add_chart(chart18, "J5")

# Δ17Oグラフの作成
chart17 = px.chart.ScatterChart()  # 散布図の設定
x17 = px.chart.Reference(sheet, min_col=2, max_col=2, min_row=2, max_row=i)  # x軸の範囲設定
y17 = px.chart.Reference(sheet, min_col=5, max_col=5, min_row=2, max_row=i)  # y軸の範囲設定
series17 = px.chart.Series(y17, x17)
series17.graphicalProperties.line.noFill = True
series17.marker.symbol = "auto"
chart17.title = "Δ17O"
chart17.x_axis.title = "1/濃度"
chart17.y_axis.title = "Δ17O"
chart17.legend = None
chart17.series.append(series17)
sheet.add_chart(chart17, "J21")

temp_name = str(input("エクセルのファイル名："))
e_name = temp_name + ".xlsx"
book.save(temp_name + ".xlsx")